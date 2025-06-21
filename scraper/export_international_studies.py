#!/usr/bin/env python3
"""Export past and current international studies to separate Excel files."""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def fetch_amed_search(keyword="国際", stage=None):
    """Fetch AMED search results."""
    base_url = "https://www.amed.go.jp/search.php"
    
    params = {
        'keyword': keyword,
        'search': 'search',
        'order_by': ''
    }
    
    if stage:
        params['stage[]'] = stage
    
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; ResearchBot/1.0)",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    }
    
    all_results = []
    page = 1
    max_pages = 10  # Limit to prevent infinite loops
    
    while page <= max_pages:
        params['page'] = page
        print(f"Fetching page {page}...")
        
        try:
            response = requests.get(base_url, params=params, headers=headers, timeout=30)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'lxml')
            
            # Find all links to recruitment pages
            links = soup.find_all('a', href=re.compile(r'/koubo/'))
            
            if not links:
                break
            
            for link in links:
                href = link.get('href', '')
                title = link.get_text(strip=True)
                
                if len(title) > 10 and '令和' in title:
                    # Extract additional info from parent elements
                    parent = link.parent
                    full_text = parent.get_text() if parent else ""
                    
                    # Extract deadline
                    deadline = None
                    deadline_patterns = [
                        r"締切[：:]\s*令和(\d+)年(\d{1,2})月(\d{1,2})日",
                        r"令和(\d+)年(\d{1,2})月(\d{1,2})日.*締切"
                    ]
                    
                    for pattern in deadline_patterns:
                        match = re.search(pattern, full_text)
                        if match:
                            year = 2018 + int(match.group(1))
                            month = int(match.group(2))
                            day = int(match.group(3))
                            deadline = f"{year}/{month}/{day}"
                            break
                    
                    # Extract program type
                    program_type = "不明"
                    if "ASPIRE" in title:
                        program_type = "ASPIRE"
                    elif "SICORP" in title:
                        program_type = "SICORP"
                    elif "e-ASIA" in title:
                        program_type = "e-ASIA"
                    elif "Interstellar" in title:
                        program_type = "Interstellar Initiative"
                    elif "SATREPS" in title:
                        program_type = "SATREPS"
                    elif "日米" in title:
                        program_type = "日米医学協力"
                    
                    # Extract countries
                    countries = []
                    country_patterns = {
                        "日・カナダ": ["日本", "カナダ"],
                        "日・スイス": ["日本", "スイス"],
                        "日・英国": ["日本", "英国"],
                        "日・フランス": ["日本", "フランス"],
                        "日・ドイツ": ["日本", "ドイツ"],
                        "日・オーストラリア": ["日本", "オーストラリア"],
                        "日米": ["日本", "アメリカ"],
                        "日・シンガポール": ["日本", "シンガポール"],
                        "日・南アフリカ": ["日本", "南アフリカ"]
                    }
                    
                    for pattern, country_list in country_patterns.items():
                        if pattern in title:
                            countries = country_list
                            break
                    
                    result = {
                        'タイトル': title,
                        'URL': f"https://www.amed.go.jp{href}" if href.startswith('/') else href,
                        '締切日': deadline,
                        'プログラム種別': program_type,
                        '対象国': ', '.join(countries) if countries else "複数国/不明",
                        '取得日時': datetime.now().strftime('%Y/%m/%d %H:%M')
                    }
                    
                    all_results.append(result)
            
            # Check if there's a next page
            next_link = soup.find('a', text=re.compile(r'次'))
            if not next_link:
                break
                
            page += 1
            
        except Exception as e:
            print(f"Error on page {page}: {e}")
            break
    
    return all_results

def create_excel_with_formatting(df, filename, title):
    """Create Excel file with formatting."""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "国際研究一覧"
    
    # Add title
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add summary
    ws['A3'] = f"総件数: {len(df)}件"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"データ取得日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
    
    # Add headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 7):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Adjust column widths
    column_widths = {
        'A': 80,  # タイトル
        'B': 50,  # URL
        'C': 15,  # 締切日
        'D': 20,  # プログラム種別
        'E': 25,  # 対象国
        'F': 20   # 取得日時
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add summary sheet
    ws2 = wb.create_sheet("サマリー")
    
    # Program type summary
    program_summary = df['プログラム種別'].value_counts()
    ws2['A1'] = "プログラム種別別集計"
    ws2['A1'].font = Font(bold=True, size=14)
    
    row = 3
    for program, count in program_summary.items():
        ws2[f'A{row}'] = program
        ws2[f'B{row}'] = count
        row += 1
    
    # Country summary
    ws2['D1'] = "対象国別集計"
    ws2['D1'].font = Font(bold=True, size=14)
    
    country_summary = df['対象国'].value_counts()
    row = 3
    for country, count in country_summary.items():
        ws2[f'D{row}'] = country
        ws2[f'E{row}'] = count
        row += 1
    
    # Save
    wb.save(filename)
    print(f"✅ Saved: {filename}")

def main():
    """Main execution function."""
    print("🔍 AMED国際研究データ取得開始...")
    
    # 1. Fetch current international studies
    print("\n📋 現在公募中の国際研究を取得中...")
    current_studies = fetch_amed_search(keyword="国際", stage="現在公募中")
    
    if current_studies:
        df_current = pd.DataFrame(current_studies)
        filename_current = f"data/AMED_国際研究_現在公募中_{datetime.now().strftime('%Y%m%d')}.xlsx"
        create_excel_with_formatting(
            df_current, 
            filename_current,
            "AMED 国際研究公募情報（現在公募中）"
        )
        print(f"現在公募中: {len(current_studies)}件")
    else:
        print("現在公募中の国際研究が見つかりませんでした")
    
    # 2. Fetch all international studies (including past)
    print("\n📋 全ての国際研究（過去含む）を取得中...")
    all_studies = fetch_amed_search(keyword="国際", stage=None)
    
    if all_studies:
        # Separate past studies
        past_studies = []
        for study in all_studies:
            # Check if it's in current studies
            is_current = any(
                current['タイトル'] == study['タイトル'] 
                for current in current_studies
            )
            if not is_current:
                past_studies.append(study)
        
        if past_studies:
            df_past = pd.DataFrame(past_studies)
            filename_past = f"data/AMED_国際研究_過去_{datetime.now().strftime('%Y%m%d')}.xlsx"
            create_excel_with_formatting(
                df_past,
                filename_past,
                "AMED 国際研究公募情報（過去）"
            )
            print(f"過去の公募: {len(past_studies)}件")
    
    print("\n✅ 完了！")
    print(f"📊 合計: {len(all_studies)}件の国際研究公募情報を取得")

if __name__ == "__main__":
    main()