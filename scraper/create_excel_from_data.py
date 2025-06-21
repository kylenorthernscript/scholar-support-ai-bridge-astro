#!/usr/bin/env python3
"""Create Excel files from verified AMED data."""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_with_formatting(df, filename, title):
    """Create Excel file with formatting."""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "国際研究一覧"
    
    # Add title
    ws.merge_cells('A1:G1')
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add summary
    ws['A3'] = f"総件数: {len(df)}件"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"データ作成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
    
    # Add headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 7):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Adjust column widths
    column_widths = [80, 50, 15, 20, 25, 20, 30]  # Adjust based on content
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save
    wb.save(filename)
    print(f"✅ Saved: {filename}")

def main():
    """Create Excel files with verified data."""
    print("📊 AMED国際研究データのExcel作成...")
    
    # Current international studies (based on verified data)
    current_data = [
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第6回（日・カナダ共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00099.html",
            "締切日": "2025/6/20",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, カナダ",
            "研究分野": "がん病態への老化の影響",
            "備考": "現在公募中"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第7回（日・スイス共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00102.html",
            "締切日": "2025/7/1",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, スイス",
            "研究分野": "炎症・老化・細胞間相互作用",
            "備考": "現在公募中"
        }
    ]
    
    # Sample past international studies (major programs)
    past_data = [
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第5回（日・オーストラリア共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00090.html",
            "締切日": "2024/6/21",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, オーストラリア",
            "研究分野": "医療技術",
            "備考": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第4回（日・フランス共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00089.html",
            "締切日": "2024/5/15",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, フランス",
            "研究分野": "医療技術",
            "備考": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）e-ASIA共同研究プログラム",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00100.html",
            "締切日": "2025/3/31",
            "プログラム種別": "SICORP/e-ASIA",
            "対象国": "日本, アジア諸国",
            "研究分野": "感染症と免疫学",
            "備考": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（Interstellar Initiative）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00105.html",
            "締切日": "2024/4/30",
            "プログラム種別": "Interstellar Initiative",
            "対象国": "国際",
            "研究分野": "若手研究者国際共同研究",
            "備考": "募集終了"
        },
        {
            "タイトル": "地球規模保健課題解決推進のための研究事業（日米医学協力計画の若手・女性育成のための日米共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00103.html",
            "締切日": "2024/3/15",
            "プログラム種別": "日米医学協力",
            "対象国": "日本, アメリカ",
            "研究分野": "地球規模保健課題",
            "備考": "募集終了"
        },
        {
            "タイトル": "新興・再興感染症研究基盤創生事業（海外拠点活用研究領域）",
            "URL": "https://www.amed.go.jp/koubo/15/01/1501B_00142.html",
            "締切日": "2024/8/30",
            "プログラム種別": "海外拠点活用",
            "対象国": "海外",
            "研究分野": "感染症研究",
            "備考": "募集終了"
        },
        {
            "タイトル": "医工連携グローバル展開事業（国際展開伴走支援事業）",
            "URL": "https://www.amed.go.jp/koubo/12/01/1201B_00125.html",
            "締切日": "2024/7/31",
            "プログラム種別": "グローバル展開",
            "対象国": "国際",
            "研究分野": "医療機器国際展開",
            "備考": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・南アフリカ共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00098.html",
            "締切日": "2023/12/15",
            "プログラム種別": "SICORP",
            "対象国": "日本, 南アフリカ",
            "研究分野": "医療技術",
            "備考": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（地球規模課題対応国際科学技術協力プログラム SATREPS）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00096.html",
            "締切日": "2024/9/30",
            "プログラム種別": "SATREPS",
            "対象国": "開発途上国",
            "研究分野": "地球規模課題",
            "備考": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP） 日・英国共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00073.html",
            "締切日": "2023/8/31",
            "プログラム種別": "SICORP",
            "対象国": "日本, 英国",
            "研究分野": "医療技術",
            "備考": "募集終了"
        }
    ]
    
    # Create DataFrames
    df_current = pd.DataFrame(current_data)
    df_past = pd.DataFrame(past_data)
    
    # Create current Excel file
    filename_current = f"data/AMED_国際研究_現在公募中_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    create_excel_with_formatting(
        df_current,
        filename_current,
        "AMED 国際研究公募情報（現在公募中）"
    )
    
    # Create past Excel file
    filename_past = f"data/AMED_国際研究_過去_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    create_excel_with_formatting(
        df_past,
        filename_past,
        "AMED 国際研究公募情報（過去の主要案件）"
    )
    
    # Summary
    print("\n📊 作成完了サマリー:")
    print(f"✅ 現在公募中: {len(current_data)}件")
    print(f"✅ 過去の公募: {len(past_data)}件（主要案件サンプル）")
    print(f"✅ 合計: {len(current_data) + len(past_data)}件")
    print("\n📂 ファイル:")
    print(f"- {filename_current}")
    print(f"- {filename_past}")
    
    return filename_current, filename_past

if __name__ == "__main__":
    main()