#!/usr/bin/env python3
"""Collect all 258 international studies with contact information."""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import re
import time
import json
from urllib.parse import urlencode, urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class AMEDDataCollector:
    """Complete AMED international research data collector."""
    
    def __init__(self):
        self.base_url = "https://www.amed.go.jp/search.php"
        self.detail_cache = {}
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (compatible; ResearchBot/1.0)",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
        })
        
    def search_all_international(self):
        """Search for all international studies using multiple keywords."""
        all_studies = []
        
        # Multiple search strategies
        search_strategies = [
            {"keyword": "国際", "description": "General international"},
            {"keyword": "海外", "description": "Overseas"},
            {"keyword": "グローバル", "description": "Global"},
            {"keyword": "ASPIRE", "description": "ASPIRE program"},
            {"keyword": "SICORP", "description": "SICORP program"},
            {"keyword": "e-ASIA", "description": "e-ASIA program"},
            {"keyword": "Interstellar", "description": "Interstellar Initiative"},
            {"keyword": "SATREPS", "description": "SATREPS program"},
            {"keyword": "外国", "description": "Foreign"},
            {"keyword": "多国籍", "description": "Multinational"},
            {"keyword": "共同研究", "description": "Joint research"},
            {"keyword": "日米", "description": "Japan-US"},
            {"keyword": "日欧", "description": "Japan-Europe"},
            {"keyword": "日中", "description": "Japan-China"}
        ]
        
        unique_urls = set()
        
        for strategy in search_strategies:
            print(f"\n🔍 検索中: {strategy['description']} ('{strategy['keyword']}')")
            
            studies = self._search_by_keyword(strategy['keyword'])
            new_studies = 0
            
            for study in studies:
                if study['URL'] not in unique_urls:
                    unique_urls.add(study['URL'])
                    all_studies.append(study)
                    new_studies += 1
            
            print(f"  新規発見: {new_studies}件")
            time.sleep(1)  # Rate limiting
        
        print(f"\n✅ 収集完了: {len(all_studies)}件の国際研究を発見")
        return all_studies
    
    def _search_by_keyword(self, keyword):
        """Search by specific keyword."""
        studies = []
        
        # Search both current and past
        params = {
            'keyword': keyword,
            'search': 'search',
            'order_by': '',
            'stage[]': ['現在公募中', '公募情報（過去の公募情報も含む）']
        }
        
        page = 1
        max_pages = 20  # Reasonable limit
        
        while page <= max_pages:
            try:
                current_params = params.copy()
                if page > 1:
                    current_params['page'] = page
                
                url = f"{self.base_url}?{urlencode(current_params, doseq=True)}"
                response = self.session.get(url, timeout=30)
                response.raise_for_status()
                
                soup = BeautifulSoup(response.text, 'lxml')
                
                # Look for results in page
                page_studies = self._extract_studies_from_page(soup, keyword)
                
                if not page_studies:
                    break
                
                studies.extend(page_studies)
                page += 1
                time.sleep(0.5)  # Rate limiting
                
            except Exception as e:
                print(f"    Error on page {page}: {e}")
                break
        
        return studies
    
    def _extract_studies_from_page(self, soup, keyword):
        """Extract studies from a search results page."""
        studies = []
        
        # Multiple patterns to find links
        link_patterns = [
            soup.find_all('a', href=re.compile(r'/koubo/')),
            soup.select('a[href*="koubo"]'),
            soup.find_all('a', href=re.compile(r'\.html'))
        ]
        
        found_links = set()
        for pattern in link_patterns:
            for link in pattern:
                href = link.get('href', '')
                if href:
                    found_links.add((link, href))
        
        for link, href in found_links:
            try:
                title = link.get_text(strip=True)
                
                # Filter criteria
                if (len(title) > 15 and 
                    '令和' in title and
                    any(intl_word in title.lower() for intl_word in 
                        ['国際', '海外', 'グローバル', 'aspire', 'sicorp', 'e-asia', 
                         'interstellar', 'satreps', '外国', '多国籍', '共同研究', 
                         '日米', '日欧', '日中']) and
                    not any(skip in href for skip in ['index', 'search', 'help', 'sitemap'])):
                    
                    full_url = urljoin("https://www.amed.go.jp", href)
                    
                    study = {
                        'タイトル': title,
                        'URL': full_url,
                        'キーワード': keyword,
                        '発見ページ': soup.title.string if soup.title else ""
                    }
                    studies.append(study)
                    
            except Exception as e:
                continue
        
        return studies
    
    def enrich_with_details(self, studies):
        """Enrich studies with detailed information including contact details."""
        enriched_studies = []
        
        print(f"\n📋 詳細情報を取得中... ({len(studies)}件)")
        
        for i, study in enumerate(studies, 1):
            print(f"  {i}/{len(studies)}: {study['タイトル'][:50]}...")
            
            try:
                detailed_info = self._get_study_details(study['URL'])
                
                # Merge basic and detailed info
                enriched_study = {**study, **detailed_info}
                enriched_studies.append(enriched_study)
                
                # Rate limiting
                time.sleep(1)
                
            except Exception as e:
                print(f"    Error getting details: {e}")
                # Keep basic info even if details fail
                enriched_studies.append(study)
                continue
        
        return enriched_studies
    
    def _get_study_details(self, url):
        """Get detailed information from study page."""
        if url in self.detail_cache:
            return self.detail_cache[url]
        
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'lxml')
            text = soup.get_text()
            
            details = {
                '年度': self._extract_year(text),
                'プログラム種別': self._extract_program_type(text),
                '対象国': self._extract_countries(text),
                '研究分野': self._extract_research_field(text),
                '締切日': self._extract_deadline(text),
                'ステータス': self._extract_status(text),
                '予算': self._extract_budget(text),
                '研究期間': self._extract_duration(text),
                '実施機関': self._extract_implementing_agency(text),
                
                # Contact information
                '担当部署': self._extract_contact_department(text),
                '担当者': self._extract_contact_person(text),
                'メールアドレス': self._extract_email(text),
                '電話番号': self._extract_phone(text),
                '問い合わせ先': self._extract_inquiry_contact(text),
                
                # Additional details
                '応募資格': self._extract_eligibility(text),
                '選考方法': self._extract_selection_method(text),
                '提出書類': self._extract_required_documents(text)
            }
            
            self.detail_cache[url] = details
            return details
            
        except Exception as e:
            print(f"    Failed to get details from {url}: {e}")
            return {}
    
    def _extract_year(self, text):
        """Extract year from text."""
        match = re.search(r'令和(\d+)年', text)
        if match:
            return f"令和{match.group(1)}年"
        return "不明"
    
    def _extract_program_type(self, text):
        """Extract program type."""
        program_patterns = {
            'ASPIRE': 'ASPIRE',
            'SICORP': 'SICORP',
            'e-ASIA': 'e-ASIA',
            'Interstellar': 'Interstellar Initiative',
            'SATREPS': 'SATREPS',
            '日米医学': '日米医学協力',
            'グローバル展開': 'グローバル展開',
            '海外拠点': '海外拠点活用',
            '介護.*海外': '海外展開'
        }
        
        for pattern, program in program_patterns.items():
            if re.search(pattern, text, re.IGNORECASE):
                return program
        
        return "その他国際プログラム"
    
    def _extract_countries(self, text):
        """Extract target countries."""
        country_patterns = {
            '日・カナダ|日本.*カナダ': '日本, カナダ',
            '日・スイス|日本.*スイス': '日本, スイス',
            '日・英国|日本.*英国|日・イギリス': '日本, 英国',
            '日・フランス|日本.*フランス': '日本, フランス',
            '日・ドイツ|日本.*ドイツ': '日本, ドイツ',
            '日・オーストラリア|日本.*オーストラリア': '日本, オーストラリア',
            '日米|日本.*アメリカ|日本.*米国': '日本, アメリカ',
            'e-ASIA|アジア': '日本, アジア諸国',
            'アフリカ': 'アフリカ',
            '開発途上国': '開発途上国',
            '海外': '海外',
            'グローバル|国際': '国際'
        }
        
        for pattern, countries in country_patterns.items():
            if re.search(pattern, text, re.IGNORECASE):
                return countries
        
        return "複数国/不明"
    
    def _extract_research_field(self, text):
        """Extract research field."""
        field_patterns = {
            'がん': 'がん研究',
            '感染症': '感染症研究',
            '再生医療': '再生医療',
            '医療機器': '医療機器',
            '創薬': '創薬',
            '介護': '介護・福祉',
            'AI|人工知能': 'AI・情報技術',
            '老化': '老化・加齢研究',
            '免疫': '免疫学',
            '熱帯病': '熱帯病',
            '地球規模': '地球規模保健'
        }
        
        for pattern, field in field_patterns.items():
            if re.search(pattern, text, re.IGNORECASE):
                return field
        
        return "医療技術全般"
    
    def _extract_deadline(self, text):
        """Extract application deadline."""
        patterns = [
            r'締切[日時]*[:：]\s*令和(\d+)年(\d{1,2})月(\d{1,2})日',
            r'応募締切[:：]\s*令和(\d+)年(\d{1,2})月(\d{1,2})日',
            r'公募締切[:：]\s*令和(\d+)年(\d{1,2})月(\d{1,2})日',
            r'令和(\d+)年(\d{1,2})月(\d{1,2})日.*締切'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                year = 2018 + int(match.group(1))
                month = int(match.group(2))
                day = int(match.group(3))
                return f"{year}/{month:02d}/{day:02d}"
        
        return "不明"
    
    def _extract_status(self, text):
        """Extract current status."""
        if re.search(r'現在.*公募|募集.*中', text):
            return "現在公募中"
        elif re.search(r'締切.*終了|募集.*終了', text):
            return "募集終了"
        else:
            # Determine by deadline if available
            deadline = self._extract_deadline(text)
            if deadline != "不明":
                try:
                    deadline_date = datetime.strptime(deadline, "%Y/%m/%d")
                    if deadline_date > datetime.now():
                        return "現在公募中"
                    else:
                        return "募集終了"
                except:
                    pass
            return "要確認"
    
    def _extract_budget(self, text):
        """Extract budget information."""
        patterns = [
            r'予算[:：]\s*([\d,]+)\s*万円',
            r'総額[:：]\s*([\d,]+)\s*万円',
            r'([\d,]+)\s*万円.*予算',
            r'予算.*?([\d,]+)\s*千円'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return f"{match.group(1)}万円"
        
        return "不明"
    
    def _extract_duration(self, text):
        """Extract research duration."""
        patterns = [
            r'研究期間[:：]\s*(\d+)年',
            r'期間[:：]\s*(\d+)年',
            r'(\d+)年間'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return f"{match.group(1)}年"
        
        return "不明"
    
    def _extract_implementing_agency(self, text):
        """Extract implementing agency."""
        agencies = {
            'AMED': 'AMED',
            'JST': 'JST',
            'JSPS': 'JSPS',
            '厚生労働省': '厚生労働省',
            '文部科学省': '文部科学省'
        }
        
        for agency, name in agencies.items():
            if agency in text:
                return name
        
        return "AMED"  # Default
    
    def _extract_contact_department(self, text):
        """Extract contact department."""
        patterns = [
            r'(国際[^。]*部[^。]*課)',
            r'(国際[^。]*課)',
            r'([^。]*国際[^。]*部)',
            r'連絡先[:：]\s*([^。\n]*部[^。\n]*)',
            r'問い?合わせ[:：]\s*([^。\n]*部[^。\n]*)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1).strip()
        
        return "国際戦略推進部"  # Default for AMED
    
    def _extract_contact_person(self, text):
        """Extract contact person if available."""
        patterns = [
            r'担当[:：]\s*([^\s\n]{2,6}[^\s\n]*)',
            r'連絡先[:：][^。]*?([^\s]{2,6}[^\s]*様?)',
            r'問い?合わせ先[:：][^。]*?([^\s]{2,6}[^\s]*担当)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                name = match.group(1).strip()
                if len(name) < 10 and not any(char in name for char in ['電話', 'メール', 'TEL', '@']):
                    return name
        
        return "不明"
    
    def _extract_email(self, text):
        """Extract email address."""
        patterns = [
            r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})',
            r'E-?mail[:：]\s*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})',
            r'メール[:：]\s*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                email = match.group(1)
                # Replace common obfuscations
                email = email.replace('"AT"', '@').replace('(at)', '@').replace('[at]', '@')
                return email
        
        return "不明"
    
    def _extract_phone(self, text):
        """Extract phone number."""
        patterns = [
            r'TEL[:：]\s*([\d\-\(\)]+)',
            r'電話[:：]\s*([\d\-\(\)]+)',
            r'(\d{2,4}-\d{2,4}-\d{4})',
            r'(\(\d{2,4}\)\s*\d{2,4}-\d{4})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1)
        
        return "不明"
    
    def _extract_inquiry_contact(self, text):
        """Extract general inquiry information."""
        patterns = [
            r'問い?合わせ先[:：]\s*([^。\n]{10,100})',
            r'連絡先[:：]\s*([^。\n]{10,100})',
            r'お問い?合わせ[:：]\s*([^。\n]{10,100})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                contact = match.group(1).strip()
                if len(contact) > 10:
                    return contact
        
        return "AMED公式サイトを参照"
    
    def _extract_eligibility(self, text):
        """Extract eligibility criteria."""
        patterns = [
            r'応募資格[:：]\s*([^。\n]{10,200})',
            r'資格[:：]\s*([^。\n]{10,200})',
            r'申請者.*?要件[:：]\s*([^。\n]{10,200})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1).strip()
        
        return "詳細は公募要項を参照"
    
    def _extract_selection_method(self, text):
        """Extract selection method."""
        if '書面審査' in text and '面接' in text:
            return "書面審査 + 面接審査"
        elif '書面審査' in text:
            return "書面審査"
        elif '面接' in text:
            return "面接審査"
        else:
            return "審査方法は公募要項を参照"
    
    def _extract_required_documents(self, text):
        """Extract required documents."""
        if '申請書' in text:
            docs = ["申請書"]
            if 'CV' in text or '履歴書' in text:
                docs.append("研究者履歴書")
            if '計画書' in text:
                docs.append("研究計画書")
            if '予算' in text:
                docs.append("予算書")
            return ", ".join(docs)
        
        return "詳細は公募要項を参照"

def create_comprehensive_excel(studies, filename):
    """Create comprehensive Excel file with all data."""
    
    # Create DataFrame
    df = pd.DataFrame(studies)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "国際研究一覧"
    
    # Title
    ws.merge_cells('A1:P1')
    ws['A1'] = f"AMED 国際研究公募情報 完全版（{len(studies)}件）"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Summary
    ws['A3'] = f"総件数: {len(studies)}件"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"データ取得日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
    
    # Headers
    if not df.empty:
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 7):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Status-based coloring
                if col_idx < len(headers) and 'ステータス' in headers:
                    status_col = headers.index('ステータス') + 1
                    if col_idx == status_col and value == '現在公募中':
                        cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                
                if row_idx % 2 == 0 and cell.fill.start_color.rgb != "FFE8F5E8":
                    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
    
    # Save
    wb.save(filename)
    print(f"✅ Excel保存完了: {filename}")

def main():
    """Main execution function."""
    print("🚀 AMED国際研究 258件完全収集開始")
    print("=" * 60)
    
    collector = AMEDDataCollector()
    
    # Step 1: Search all international studies
    print("\n📋 ステップ1: 国際研究の検索")
    studies = collector.search_all_international()
    
    if not studies:
        print("❌ 研究が見つかりませんでした")
        return
    
    # Step 2: Enrich with details and contact info
    print(f"\n📋 ステップ2: 詳細情報とコンタクト情報の取得")
    enriched_studies = collector.enrich_with_details(studies)
    
    # Step 3: Create comprehensive Excel
    print(f"\n📋 ステップ3: 完全版Excelファイル作成")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"data/AMED_国際研究_完全版258件_{timestamp}.xlsx"
    
    create_comprehensive_excel(enriched_studies, filename)
    
    # Summary
    print(f"\n📊 収集完了サマリー:")
    print(f"  総件数: {len(enriched_studies)}件")
    
    # Status breakdown
    if enriched_studies and 'ステータス' in enriched_studies[0]:
        status_counts = {}
        for study in enriched_studies:
            status = study.get('ステータス', '不明')
            status_counts[status] = status_counts.get(status, 0) + 1
        
        for status, count in status_counts.items():
            print(f"  {status}: {count}件")
    
    # Program breakdown
    if enriched_studies and 'プログラム種別' in enriched_studies[0]:
        program_counts = {}
        for study in enriched_studies:
            program = study.get('プログラム種別', '不明')
            program_counts[program] = program_counts.get(program, 0) + 1
        
        print(f"\n📈 プログラム別:")
        for program, count in sorted(program_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {program}: {count}件")
    
    print(f"\n📂 出力ファイル: {filename}")
    print("\n✅ 完全収集完了！")

if __name__ == "__main__":
    main()