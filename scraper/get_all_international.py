#!/usr/bin/env python3
"""Get all 258 international studies from AMED."""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import re
import time
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def get_all_international_studies():
    """Get all international studies by scraping search results."""
    base_url = "https://www.amed.go.jp/search.php"
    
    # Search for all international studies (past + current)
    params = {
        'keyword': '国際',
        'search': 'search',
        'order_by': '',
        'stage[]': ['現在公募中', '公募情報（過去の公募情報も含む）']
    }
    
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; ResearchBot/1.0)",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    }
    
    all_studies = []
    page = 1
    max_pages = 30  # Should be enough for 258 results (10 per page = 26 pages)
    
    print("🔍 AMED国際研究の全件取得を開始...")
    
    while page <= max_pages:
        print(f"📄 ページ {page} を取得中...")
        
        try:
            # Add page parameter
            current_params = params.copy()
            if page > 1:
                current_params['page'] = page
            
            # Build URL
            url = f"{base_url}?{urlencode(current_params, doseq=True)}"
            
            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'lxml')
            
            # Check total count on first page
            if page == 1:
                count_text = soup.get_text()
                match = re.search(r'検索結果.*?(\d+)件中', count_text)
                if match:
                    total_count = int(match.group(1))
                    print(f"📊 総件数: {total_count}件")
                    max_pages = min(max_pages, (total_count // 10) + 2)
            
            # Find result items - try multiple selectors
            result_items = []
            
            # Try different possible selectors for search results
            selectors = [
                'div.searchResult',
                'div.searchResultItem', 
                'li.searchResult',
                'tr.searchResult',
                'div.result',
                'li.result'
            ]
            
            for selector in selectors:
                items = soup.select(selector)
                if items:
                    result_items = items
                    print(f"   見つかったセレクタ: {selector} ({len(items)}件)")
                    break
            
            # If no structured results, look for links
            if not result_items:
                print("   構造化結果が見つからない - リンクを直接検索")
                links = soup.find_all('a', href=re.compile(r'/koubo/'))
                
                current_page_studies = []
                for link in links:
                    href = link.get('href', '')
                    title = link.get_text(strip=True)
                    
                    # Filter for valid recruitment links
                    if (title and len(title) > 10 and 
                        '令和' in title and 
                        '国際' in title and
                        not any(skip in href for skip in ['index', 'search', 'help'])):
                        
                        # Avoid duplicates
                        if not any(study['タイトル'] == title for study in all_studies):
                            study_data = parse_study_from_link(link, title, href)
                            if study_data:
                                current_page_studies.append(study_data)
                
                if current_page_studies:
                    all_studies.extend(current_page_studies)
                    print(f"   このページで{len(current_page_studies)}件取得")
                else:
                    print(f"   このページでは新しい結果が見つかりませんでした")
                    break
            
            else:
                # Parse structured result items
                current_page_studies = []
                for item in result_items:
                    study_data = parse_study_from_item(item)
                    if study_data:
                        current_page_studies.append(study_data)
                
                if current_page_studies:
                    all_studies.extend(current_page_studies)
                    print(f"   このページで{len(current_page_studies)}件取得")
                else:
                    break
            
            # Rate limiting
            time.sleep(1)
            page += 1
            
        except Exception as e:
            print(f"❌ ページ {page} でエラー: {e}")
            break
    
    print(f"\n✅ 取得完了: {len(all_studies)}件")
    return all_studies

def parse_study_from_link(link, title, href):
    """Parse study data from a link element."""
    try:
        # Get surrounding context
        parent = link.parent
        if parent:
            context = parent.get_text()
        else:
            context = title
        
        # Extract year from title
        year_match = re.search(r'令和(\d+)年', title)
        year = None
        if year_match:
            year = 2018 + int(year_match.group(1))
        
        # Determine program type
        program_type = "不明"
        if "ASPIRE" in title:
            program_type = "ASPIRE"
        elif "SICORP" in title:
            program_type = "SICORP"  
        elif "e-ASIA" in title:
            program_type = "e-ASIA"
        elif "Interstellar" in title:
            program_type = "Interstellar Initiative"
        elif "SATREPS" in title:
            program_type = "SATREPS"
        elif "日米" in title:
            program_type = "日米医学協力"
        elif "海外" in title:
            program_type = "海外拠点活用"
        elif "グローバル" in title:
            program_type = "グローバル展開"
        
        # Extract countries
        countries = extract_countries_from_title(title)
        
        # Determine status
        status = "募集終了" if year and year < 2025 else "確認要"
        
        return {
            'タイトル': title,
            'URL': f"https://www.amed.go.jp{href}" if href.startswith('/') else href,
            '年度': f"令和{year-2018}年" if year else "不明",
            'プログラム種別': program_type,
            '対象国': countries,
            'ステータス': status,
            '取得日時': datetime.now().strftime('%Y/%m/%d %H:%M')
        }
        
    except Exception as e:
        print(f"Warning: パース中にエラー: {e}")
        return None

def parse_study_from_item(item):
    """Parse study data from a structured result item."""
    try:
        # Find link within item
        link = item.find('a', href=re.compile(r'/koubo/'))
        if not link:
            return None
        
        title = link.get_text(strip=True)
        href = link.get('href', '')
        
        return parse_study_from_link(link, title, href)
        
    except Exception as e:
        print(f"Warning: 構造化アイテムのパース中にエラー: {e}")
        return None

def extract_countries_from_title(title):
    """Extract countries from title."""
    country_patterns = {
        "日・カナダ": "日本, カナダ",
        "日・スイス": "日本, スイス", 
        "日・英国": "日本, 英国",
        "日・フランス": "日本, フランス",
        "日・ドイツ": "日本, ドイツ",
        "日・オーストラリア": "日本, オーストラリア",
        "日・シンガポール": "日本, シンガポール",
        "日・南アフリカ": "日本, 南アフリカ",
        "日・リトアニア": "日本, リトアニア",
        "日・北欧": "日本, 北欧諸国",
        "日米": "日本, アメリカ",
        "日欧": "日本, 欧州",
        "日中": "日本, 中国",
        "日韓": "日本, 韓国",
        "e-ASIA": "日本, アジア諸国",
        "海外": "海外",
        "グローバル": "国際",
        "アフリカ": "アフリカ"
    }
    
    for pattern, countries in country_patterns.items():
        if pattern in title:
            return countries
    
    return "複数国/不明"

def create_excel_with_formatting(df, filename, title):
    """Create Excel file with formatting."""
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "国際研究一覧"
    
    # Add title
    ws.merge_cells('A1:G1')
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add summary
    ws['A3'] = f"総件数: {len(df)}件"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"データ取得日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
    
    # Add headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 7):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Adjust column widths
    column_widths = [80, 50, 12, 25, 25, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Add summary sheet
    ws2 = wb.create_sheet("プログラム別集計")
    
    # Program type summary
    program_summary = df['プログラム種別'].value_counts()
    ws2['A1'] = "プログラム種別別件数"
    ws2['A1'].font = Font(bold=True, size=14)
    
    row = 3
    for program, count in program_summary.items():
        ws2[f'A{row}'] = program
        ws2[f'B{row}'] = count
        row += 1
    
    # Save
    wb.save(filename)
    print(f"✅ Excel保存完了: {filename}")

def main():
    """Main execution function."""
    # Get all international studies
    all_studies = get_all_international_studies()
    
    if not all_studies:
        print("❌ データが取得できませんでした")
        return
    
    # Create DataFrame
    df = pd.DataFrame(all_studies)
    
    # Remove duplicates based on title
    df = df.drop_duplicates(subset=['タイトル'])
    
    print(f"📊 重複除去後: {len(df)}件")
    
    # Create Excel file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"data/AMED_国際研究_全件_{timestamp}.xlsx"
    
    create_excel_with_formatting(
        df,
        filename, 
        f"AMED 国際研究公募情報 全{len(df)}件"
    )
    
    # Show summary
    print(f"\n📈 プログラム別集計:")
    program_counts = df['プログラム種別'].value_counts()
    for program, count in program_counts.items():
        print(f"  {program}: {count}件")
    
    print(f"\n📂 保存ファイル: {filename}")

if __name__ == "__main__":
    main()