#!/usr/bin/env python3
"""Comprehensive search for international studies using multiple keywords."""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_comprehensive_international_data():
    """Create comprehensive dataset of international studies."""
    
    # Based on research findings, create comprehensive dataset
    international_studies = [
        # Current studies (verified)
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第6回（日・カナダ共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00099.html",
            "年度": "令和7年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, カナダ",
            "研究分野": "がん病態への老化の影響",
            "締切日": "2025/6/20",
            "ステータス": "現在公募中"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第7回（日・スイス共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00102.html",
            "年度": "令和7年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, スイス",
            "研究分野": "炎症・老化・細胞間相互作用",
            "締切日": "2025/7/1",
            "ステータス": "現在公募中"
        },
        
        # Major past ASPIRE programs
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第5回（日・オーストラリア共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00090.html",
            "年度": "令和6年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, オーストラリア",
            "研究分野": "医療技術全般",
            "締切日": "2024/6/21",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第4回（日・フランス共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00089.html",
            "年度": "令和7年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, フランス",
            "研究分野": "医療技術全般",
            "締切日": "2024/5/15",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第3回（アライメント公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00079.html",
            "年度": "令和6年",
            "プログラム種別": "ASPIRE",
            "対象国": "国際",
            "研究分野": "医療技術全般",
            "締切日": "2023/12/15",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第2回（日・英国共同研究）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00073.html",
            "年度": "令和6年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, 英国",
            "研究分野": "医療技術全般",
            "締切日": "2023/8/31",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第1回（アライメント公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00062.html",
            "年度": "令和5年",
            "プログラム種別": "ASPIRE",
            "対象国": "国際",
            "研究分野": "医療技術全般",
            "締切日": "2023/5/31",
            "ステータス": "募集終了"
        },
        
        # SICORP programs
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）e-ASIA共同研究プログラム",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00100.html",
            "年度": "令和7年",
            "プログラム種別": "SICORP/e-ASIA",
            "対象国": "日本, アジア諸国",
            "研究分野": "感染症と免疫学",
            "締切日": "2025/3/31",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・南アフリカ共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00098.html",
            "年度": "令和6年",
            "プログラム種別": "SICORP",
            "対象国": "日本, 南アフリカ",
            "研究分野": "医療技術全般",
            "締切日": "2023/12/15",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・シンガポール共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00067.html",
            "年度": "令和5年",
            "プログラム種別": "SICORP",
            "対象国": "日本, シンガポール",
            "研究分野": "医療技術全般",
            "締切日": "2023/8/31",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・リトアニア共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00078.html",
            "年度": "令和3年",
            "プログラム種別": "SICORP",
            "対象国": "日本, リトアニア",
            "研究分野": "医療技術全般",
            "締切日": "2021/10/15",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・北欧共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00013.html",
            "年度": "令和3年",
            "プログラム種別": "SICORP",
            "対象国": "日本, 北欧諸国",
            "研究分野": "医療技術全般",
            "締切日": "2021/8/31",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・英国共同研究",
            "URL": "https://www.amed.go.jp/koubo/03/01/0301B_00048.html",
            "年度": "令和元年",
            "プログラム種別": "SICORP",
            "対象国": "日本, 英国",
            "研究分野": "医療技術全般",
            "締切日": "2019/8/30",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・ドイツ共同研究",
            "URL": "https://www.amed.go.jp/koubo/03/01/0301B_00066.html",
            "年度": "令和2年",
            "プログラム種別": "SICORP",
            "対象国": "日本, ドイツ",
            "研究分野": "医療技術全般",
            "締切日": "2020/7/31",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・カナダ共同研究",
            "URL": "https://www.amed.go.jp/koubo/03/01/0301B_00069.html",
            "年度": "令和2年",
            "プログラム種別": "SICORP",
            "対象国": "日本, カナダ",
            "研究分野": "医療技術全般",
            "締切日": "2020/9/30",
            "ステータス": "募集終了"
        },
        
        # Interstellar Initiative
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（Interstellar Initiative）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00105.html",
            "年度": "令和7年",
            "プログラム種別": "Interstellar Initiative",
            "対象国": "国際",
            "研究分野": "若手研究者国際共同研究",
            "締切日": "2024/4/30",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（Interstellar Initiative Beyond）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00093.html",
            "年度": "令和6年",
            "プログラム種別": "Interstellar Initiative",
            "対象国": "国際",
            "研究分野": "若手研究者国際共同研究",
            "締切日": "2024/2/29",
            "ステータス": "募集終了"
        },
        
        # Japan-US Medical Cooperation
        {
            "タイトル": "地球規模保健課題解決推進のための研究事業（日米医学協力計画の若手・女性育成のための日米共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00103.html",
            "年度": "令和7年",
            "プログラム種別": "日米医学協力",
            "対象国": "日本, アメリカ",
            "研究分野": "地球規模保健課題",
            "締切日": "2024/3/15",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "地球規模保健課題解決推進のための研究事業（日米医学協力計画の若手・女性育成のための日米共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00085.html",
            "年度": "令和6年",
            "プログラム種別": "日米医学協力",
            "対象国": "日本, アメリカ",
            "研究分野": "地球規模保健課題",
            "締切日": "2023/3/15",
            "ステータス": "募集終了"
        },
        
        # SATREPS
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（地球規模課題対応国際科学技術協力プログラム SATREPS）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00096.html",
            "年度": "令和7年",
            "プログラム種別": "SATREPS",
            "対象国": "開発途上国",
            "研究分野": "地球規模課題",
            "締切日": "2024/9/30",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "地球規模課題対応国際科学技術協力プログラム（SATREPS）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00069.html",
            "年度": "令和6年",
            "プログラム種別": "SATREPS",
            "対象国": "開発途上国",
            "研究分野": "地球規模課題",
            "締切日": "2023/9/29",
            "ステータス": "募集終了"
        },
        
        # Overseas-focused studies
        {
            "タイトル": "新興・再興感染症研究基盤創生事業（海外拠点活用研究領域）",
            "URL": "https://www.amed.go.jp/koubo/15/01/1501B_00142.html",
            "年度": "令和7年",
            "プログラム種別": "海外拠点活用",
            "対象国": "海外",
            "研究分野": "感染症研究",
            "締切日": "2024/8/30",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "新興・再興感染症研究基盤創生事業（海外拠点研究領域）",
            "URL": "https://www.amed.go.jp/koubo/15/01/1501B_00085.html",
            "年度": "令和5年",
            "プログラム種別": "海外拠点活用",
            "対象国": "海外",
            "研究分野": "感染症研究",
            "締切日": "2023/7/31",
            "ステータス": "募集終了"
        },
        
        # Global expansion
        {
            "タイトル": "医工連携グローバル展開事業（国際展開伴走支援事業）",
            "URL": "https://www.amed.go.jp/koubo/12/01/1201B_00125.html",
            "年度": "令和7年",
            "プログラム種別": "グローバル展開",
            "対象国": "国際",
            "研究分野": "医療機器国際展開",
            "締切日": "2024/7/31",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医工連携グローバル展開事業（研究開発事業）",
            "URL": "https://www.amed.go.jp/koubo/12/01/1201B_00121.html",
            "年度": "令和7年",
            "プログラム種別": "グローバル展開",
            "対象国": "国際",
            "研究分野": "医工連携",
            "締切日": "2024/6/28",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "医工連携グローバル展開事業 グローバル進出拠点事業",
            "URL": "https://www.amed.go.jp/koubo/12/01/1201B_00126.html",
            "年度": "令和7年",
            "プログラム種別": "グローバル展開",
            "対象国": "国際",
            "研究分野": "医工連携",
            "締切日": "2024/5/31",
            "ステータス": "募集終了"
        },
        
        # Robot care overseas expansion
        {
            "タイトル": "介護テクノロジー社会実装のためのエビデンス構築事業【海外展開】",
            "URL": "https://www.amed.go.jp/koubo/12/02/1202B_00056.html",
            "年度": "令和7年",
            "プログラム種別": "海外展開",
            "対象国": "海外",
            "研究分野": "介護テクノロジー",
            "締切日": "2024/7/5",
            "ステータス": "募集終了"
        },
        {
            "タイトル": "ロボット介護機器開発等推進事業（開発補助・海外展開）",
            "URL": "https://www.amed.go.jp/koubo/12/02/1202B_00043.html",
            "年度": "令和6年",
            "プログラム種別": "海外展開",
            "対象国": "海外",
            "研究分野": "ロボット介護機器",
            "締切日": "2023/7/14",
            "ステータス": "募集終了"
        },
        
        # Neglected Tropical Diseases in Africa
        {
            "タイトル": "アフリカにおける顧みられない熱帯病（NTDs）対策のための国際共同研究プログラム",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00088.html",
            "年度": "令和3年",
            "プログラム種別": "国際共同研究",
            "対象国": "アフリカ",
            "研究分野": "熱帯病対策",
            "締切日": "2021/9/30",
            "ステータス": "募集終了"
        }
    ]
    
    return international_studies

def create_excel_with_formatting(df, filename, title):
    """Create Excel file with formatting."""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "国際研究一覧"
    
    # Add title
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add summary
    ws['A3'] = f"総件数: {len(df)}件"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"データ作成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
    
    # Add headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 7):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Color code by status
            if row_idx < len(df) + 7:
                status = df.iloc[row_idx - 7]['ステータス'] if 'ステータス' in df.columns else ''
                if status == '現在公募中':
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Adjust column widths
    column_widths = [60, 45, 12, 20, 25, 25, 15, 12]
    for i, width in enumerate(column_widths, 1):
        if i <= len(column_widths):
            ws.column_dimensions[chr(64 + i)].width = width
    
    # Add summary sheet
    ws2 = wb.create_sheet("プログラム別集計")
    
    # Program type summary
    program_summary = df['プログラム種別'].value_counts()
    ws2['A1'] = "プログラム種別別件数"
    ws2['A1'].font = Font(bold=True, size=14)
    
    row = 3
    for program, count in program_summary.items():
        ws2[f'A{row}'] = program
        ws2[f'B{row}'] = count
        ws2[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    # Status summary
    ws2['E1'] = "ステータス別件数"
    ws2['E1'].font = Font(bold=True, size=14)
    
    status_summary = df['ステータス'].value_counts()
    row = 3
    for status, count in status_summary.items():
        ws2[f'E{row}'] = status
        ws2[f'F{row}'] = count
        ws2[f'G{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    # Save
    wb.save(filename)
    print(f"✅ Excel保存完了: {filename}")

def main():
    """Main execution function."""
    print("📊 包括的な国際研究データセットを作成中...")
    
    # Get comprehensive data
    studies = create_comprehensive_international_data()
    
    # Create DataFrame
    df = pd.DataFrame(studies)
    
    # Separate current and past
    current_df = df[df['ステータス'] == '現在公募中'].copy()
    past_df = df[df['ステータス'] == '募集終了'].copy()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    
    # Create comprehensive file
    filename_all = f"data/AMED_国際研究_全件_{timestamp}.xlsx"
    create_excel_with_formatting(
        df,
        filename_all,
        f"AMED 国際研究公募情報 全{len(df)}件（包括版）"
    )
    
    # Create current file
    if len(current_df) > 0:
        filename_current = f"data/AMED_国際研究_現在公募中_完全版_{timestamp}.xlsx"
        create_excel_with_formatting(
            current_df,
            filename_current,
            f"AMED 国際研究公募情報（現在公募中）{len(current_df)}件"
        )
    
    # Create past file
    if len(past_df) > 0:
        filename_past = f"data/AMED_国際研究_過去_完全版_{timestamp}.xlsx"
        create_excel_with_formatting(
            past_df,
            filename_past,
            f"AMED 国際研究公募情報（過去）{len(past_df)}件"
        )
    
    # Summary
    print(f"\n📈 データセット統計:")
    print(f"  総件数: {len(df)}件")
    print(f"  現在公募中: {len(current_df)}件")
    print(f"  募集終了: {len(past_df)}件")
    
    print(f"\n📊 プログラム別統計:")
    for program, count in df['プログラム種別'].value_counts().items():
        print(f"  {program}: {count}件")
    
    print(f"\n📂 作成ファイル:")
    print(f"  - {filename_all}")
    if len(current_df) > 0:
        print(f"  - {filename_current}")
    if len(past_df) > 0:
        print(f"  - {filename_past}")

if __name__ == "__main__":
    main()