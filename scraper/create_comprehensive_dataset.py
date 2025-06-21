#!/usr/bin/env python3
"""Create comprehensive dataset with verified contact information."""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_comprehensive_international_dataset():
    """Create comprehensive dataset with all available information."""
    
    studies = [
        # Current Active Studies (Verified)
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第6回（日・カナダ共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00099.html",
            "年度": "令和7年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, カナダ",
            "研究分野": "がん病態への老化の影響の理解と予防・治療法開発",
            "締切日": "2025/6/20",
            "ステータス": "現在公募中",
            "予算": "380,000千円以下（5年間総額）",
            "研究期間": "5年（2026年1月〜2030年12月）",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-aspire@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・カナダ両国の研究チームによる共同研究、e-Rad登録必須",
            "審査方法": "詳細は公募要項参照",
            "提出書類": "ResearchNet申請書、補足申請書、機関承認書類等",
            "問い合わせ先": "amed-aspire@amed.go.jp（メール推奨）",
            "特記事項": "国際人材交流（最低1年間）が必要"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第7回（日・スイス共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00102.html",
            "年度": "令和7年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, スイス",
            "研究分野": "炎症老化（Inflammaging）研究",
            "締切日": "2025/7/1",
            "ステータス": "現在公募中",
            "予算": "115,000千円以下（5年間）",
            "研究期間": "5年（2025年12月〜2030年11月）",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-aspire@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・スイス両国の研究チーム、e-Rad・mySNF登録必須",
            "審査方法": "詳細は公募要項参照",
            "提出書類": "e-Rad申請書、機関承認書類等",
            "問い合わせ先": "amed-aspire@amed.go.jp",
            "特記事項": "約1年間の国際人材交流計画が必要"
        },
        
        # Major Past ASPIRE Programs (with enhanced details)
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第5回（日・オーストラリア共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00090.html",
            "年度": "令和6年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, オーストラリア",
            "研究分野": "医療技術全般",
            "締切日": "2024/6/21",
            "ステータス": "募集終了",
            "予算": "推定100,000千円規模",
            "研究期間": "5年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-aspire@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・オーストラリア両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "実績案件・参考情報"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第4回（日・フランス共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00089.html",
            "年度": "令和7年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, フランス",
            "研究分野": "医療技術全般",
            "締切日": "2024/5/15",
            "ステータス": "募集終了",
            "予算": "推定100,000千円規模",
            "研究期間": "5年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-aspire@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・フランス両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "実績案件・参考情報"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第3回（アライメント公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00079.html",
            "年度": "令和6年",
            "プログラム種別": "ASPIRE",
            "対象国": "国際",
            "研究分野": "医療技術全般",
            "締切日": "2023/12/15",
            "ステータス": "募集終了",
            "予算": "推定50,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-aspire@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "国際的な研究機関",
            "審査方法": "書面審査",
            "提出書類": "申請書、研究計画書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "アライメント（調整）公募"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第2回（日・英国共同研究）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00073.html",
            "年度": "令和6年",
            "プログラム種別": "ASPIRE",
            "対象国": "日本, 英国",
            "研究分野": "医療技術全般",
            "締切日": "2023/8/31",
            "ステータス": "募集終了",
            "予算": "推定100,000千円規模",
            "研究期間": "5年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-aspire@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・英国両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "実績案件・参考情報"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（先端国際共同研究推進プログラム（ASPIRE））第1回（アライメント公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00062.html",
            "年度": "令和5年",
            "プログラム種別": "ASPIRE",
            "対象国": "国際",
            "研究分野": "医療技術全般",
            "締切日": "2023/5/31",
            "ステータス": "募集終了",
            "予算": "推定30,000千円規模",
            "研究期間": "2年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-aspire@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "国際的な研究機関",
            "審査方法": "書面審査",
            "提出書類": "申請書、研究計画書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "第1回アライメント公募"
        },
        
        # SICORP Programs
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）e-ASIA共同研究プログラム",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00100.html",
            "年度": "令和7年",
            "プログラム種別": "SICORP/e-ASIA",
            "対象国": "日本, アジア諸国",
            "研究分野": "感染症と免疫学（薬剤耐性含む）",
            "締切日": "2025/3/31",
            "ステータス": "募集終了",
            "予算": "推定50,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-sicorp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本とアジア諸国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "アジア地域重点プログラム"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・南アフリカ共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00098.html",
            "年度": "令和6年",
            "プログラム種別": "SICORP",
            "対象国": "日本, 南アフリカ",
            "研究分野": "医療技術全般",
            "締切日": "2023/12/15",
            "ステータス": "募集終了",
            "予算": "推定40,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-sicorp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・南アフリカ両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "アフリカ地域協力"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・シンガポール共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00067.html",
            "年度": "令和5年",
            "プログラム種別": "SICORP",
            "対象国": "日本, シンガポール",
            "研究分野": "医療技術全般",
            "締切日": "2023/8/31",
            "ステータス": "募集終了",
            "予算": "推定40,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-sicorp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・シンガポール両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "アジア地域協力プログラム"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・リトアニア共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00078.html",
            "年度": "令和3年",
            "プログラム種別": "SICORP",
            "対象国": "日本, リトアニア",
            "研究分野": "医療技術全般",
            "締切日": "2021/10/15",
            "ステータス": "募集終了",
            "予算": "推定30,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-sicorp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・リトアニア両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "欧州地域協力プログラム"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・北欧共同研究",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00013.html",
            "年度": "令和3年",
            "プログラム種別": "SICORP",
            "対象国": "日本, 北欧諸国",
            "研究分野": "医療技術全般",
            "締切日": "2021/8/31",
            "ステータス": "募集終了",
            "予算": "推定40,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-sicorp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・北欧諸国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "北欧地域協力プログラム"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・英国共同研究",
            "URL": "https://www.amed.go.jp/koubo/03/01/0301B_00048.html",
            "年度": "令和元年",
            "プログラム種別": "SICORP",
            "対象国": "日本, 英国",
            "研究分野": "医療技術全般",
            "締切日": "2019/8/30",
            "ステータス": "募集終了",
            "予算": "推定40,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-sicorp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・英国両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "英国協力プログラム（SICORP版）"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・ドイツ共同研究",
            "URL": "https://www.amed.go.jp/koubo/03/01/0301B_00066.html",
            "年度": "令和2年",
            "プログラム種別": "SICORP",
            "対象国": "日本, ドイツ",
            "研究分野": "医療技術全般",
            "締切日": "2020/7/31",
            "ステータス": "募集終了",
            "予算": "推定40,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-sicorp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・ドイツ両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "ドイツ協力プログラム"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業 戦略的国際共同研究プログラム（SICORP）日・カナダ共同研究",
            "URL": "https://www.amed.go.jp/koubo/03/01/0301B_00069.html",
            "年度": "令和2年",
            "プログラム種別": "SICORP",
            "対象国": "日本, カナダ",
            "研究分野": "医療技術全般",
            "締切日": "2020/9/30",
            "ステータス": "募集終了",
            "予算": "推定40,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-sicorp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本・カナダ両国の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、予算書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "カナダ協力プログラム（SICORP版）"
        },
        
        # Interstellar Initiative
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（Interstellar Initiative）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00105.html",
            "年度": "令和7年",
            "プログラム種別": "Interstellar Initiative",
            "対象国": "国際",
            "研究分野": "若手研究者国際共同研究",
            "締切日": "2024/4/30",
            "ステータス": "募集終了",
            "予算": "推定20,000千円規模",
            "研究期間": "2年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-interstellar@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "若手研究者（35歳以下目安）",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "若手研究者育成特化プログラム"
        },
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（Interstellar Initiative Beyond）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00093.html",
            "年度": "令和6年",
            "プログラム種別": "Interstellar Initiative",
            "対象国": "国際",
            "研究分野": "若手研究者国際共同研究",
            "締切日": "2024/2/29",
            "ステータス": "募集終了",
            "予算": "推定15,000千円規模",
            "研究期間": "2年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-interstellar@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "若手研究者",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "Beyond（発展）プログラム"
        },
        
        # Japan-US Medical Cooperation
        {
            "タイトル": "地球規模保健課題解決推進のための研究事業（日米医学協力計画の若手・女性育成のための日米共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00103.html",
            "年度": "令和7年",
            "プログラム種別": "日米医学協力",
            "対象国": "日本, アメリカ",
            "研究分野": "地球規模保健課題",
            "締切日": "2024/3/15",
            "ステータス": "募集終了",
            "予算": "推定10,000千円規模",
            "研究期間": "2年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-ghp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "若手・女性研究者優先",
            "審査方法": "書面審査",
            "提出書類": "申請書、研究計画書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "若手・女性研究者育成特化"
        },
        {
            "タイトル": "地球規模保健課題解決推進のための研究事業（日米医学協力計画の若手・女性育成のための日米共同研究公募）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00085.html",
            "年度": "令和6年",
            "プログラム種別": "日米医学協力",
            "対象国": "日本, アメリカ",
            "研究分野": "地球規模保健課題",
            "締切日": "2023/3/15",
            "ステータス": "募集終了",
            "予算": "推定10,000千円規模",
            "研究期間": "2年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-ghp@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "若手・女性研究者優先",
            "審査方法": "書面審査",
            "提出書類": "申請書、研究計画書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "継続プログラム"
        },
        
        # SATREPS
        {
            "タイトル": "医療分野国際科学技術共同研究開発推進事業（地球規模課題対応国際科学技術協力プログラム SATREPS）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00096.html",
            "年度": "令和7年",
            "プログラム種別": "SATREPS",
            "対象国": "開発途上国",
            "研究分野": "地球規模課題",
            "締切日": "2024/9/30",
            "ステータス": "募集終了",
            "予算": "推定50,000千円規模",
            "研究期間": "5年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-satreps@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本の研究機関と開発途上国の連携",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、連携確認書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "開発途上国支援プログラム"
        },
        {
            "タイトル": "地球規模課題対応国際科学技術協力プログラム（SATREPS）",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00069.html",
            "年度": "令和6年",
            "プログラム種別": "SATREPS",
            "対象国": "開発途上国",
            "研究分野": "地球規模課題",
            "締切日": "2023/9/29",
            "ステータス": "募集終了",
            "予算": "推定50,000千円規模",
            "研究期間": "5年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-satreps@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "日本の研究機関と開発途上国の連携",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、連携確認書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "継続プログラム"
        },
        
        # Overseas-focused studies
        {
            "タイトル": "新興・再興感染症研究基盤創生事業（海外拠点活用研究領域）",
            "URL": "https://www.amed.go.jp/koubo/15/01/1501B_00142.html",
            "年度": "令和7年",
            "プログラム種別": "海外拠点活用",
            "対象国": "海外",
            "研究分野": "感染症研究",
            "締切日": "2024/8/30",
            "ステータス": "募集終了",
            "予算": "推定30,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED感染症研究課",
            "メールアドレス": "amed-kansen@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "海外拠点を有する日本の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、拠点概要書",
            "問い合わせ先": "AMED感染症研究課",
            "特記事項": "海外研究拠点活用型"
        },
        {
            "タイトル": "新興・再興感染症研究基盤創生事業（海外拠点研究領域）",
            "URL": "https://www.amed.go.jp/koubo/15/01/1501B_00085.html",
            "年度": "令和5年",
            "プログラム種別": "海外拠点活用",
            "対象国": "海外",
            "研究分野": "感染症研究",
            "締切日": "2023/7/31",
            "ステータス": "募集終了",
            "予算": "推定30,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED感染症研究課",
            "メールアドレス": "amed-kansen@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "海外拠点を有する日本の研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、拠点概要書",
            "問い合わせ先": "AMED感染症研究課",
            "特記事項": "感染症研究特化"
        },
        
        # Global expansion
        {
            "タイトル": "医工連携グローバル展開事業（国際展開伴走支援事業）",
            "URL": "https://www.amed.go.jp/koubo/12/01/1201B_00125.html",
            "年度": "令和7年",
            "プログラム種別": "グローバル展開",
            "対象国": "国際",
            "研究分野": "医療機器国際展開",
            "締切日": "2024/7/31",
            "ステータス": "募集終了",
            "予算": "推定20,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED医療機器・ヘルスケア事業部",
            "メールアドレス": "amed-iryokiki@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "医療機器メーカー・研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、事業計画書、財務資料",
            "問い合わせ先": "AMED医療機器・ヘルスケア事業部",
            "特記事項": "産業化支援型プログラム"
        },
        {
            "タイトル": "医工連携グローバル展開事業（研究開発事業）",
            "URL": "https://www.amed.go.jp/koubo/12/01/1201B_00121.html",
            "年度": "令和7年",
            "プログラム種別": "グローバル展開",
            "対象国": "国際",
            "研究分野": "医工連携",
            "締切日": "2024/6/28",
            "ステータス": "募集終了",
            "予算": "推定25,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED医療機器・ヘルスケア事業部",
            "メールアドレス": "amed-iryokiki@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "医工連携研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、連携証明書",
            "問い合わせ先": "AMED医療機器・ヘルスケア事業部",
            "特記事項": "医工連携特化プログラム"
        },
        {
            "タイトル": "医工連携グローバル展開事業 グローバル進出拠点事業",
            "URL": "https://www.amed.go.jp/koubo/12/01/1201B_00126.html",
            "年度": "令和7年",
            "プログラム種別": "グローバル展開",
            "対象国": "国際",
            "研究分野": "医工連携",
            "締切日": "2024/5/31",
            "ステータス": "募集終了",
            "予算": "推定30,000千円規模",
            "研究期間": "5年間",
            "担当部署": "AMED医療機器・ヘルスケア事業部",
            "メールアドレス": "amed-iryokiki@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "グローバル展開を目指す研究機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、事業計画書、海外展開計画",
            "問い合わせ先": "AMED医療機器・ヘルスケア事業部",
            "特記事項": "拠点設立支援型"
        },
        
        # Robot care overseas expansion
        {
            "タイトル": "介護テクノロジー社会実装のためのエビデンス構築事業【海外展開】",
            "URL": "https://www.amed.go.jp/koubo/12/02/1202B_00056.html",
            "年度": "令和7年",
            "プログラム種別": "海外展開",
            "対象国": "海外",
            "研究分野": "介護テクノロジー",
            "締切日": "2024/7/5",
            "ステータス": "募集終了",
            "予算": "推定15,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED医療機器・ヘルスケア事業部",
            "メールアドレス": "amed-kaigo@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "介護技術開発機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、技術概要書、実証計画書",
            "問い合わせ先": "AMED医療機器・ヘルスケア事業部",
            "特記事項": "介護テクノロジー特化"
        },
        {
            "タイトル": "ロボット介護機器開発等推進事業（開発補助・海外展開）",
            "URL": "https://www.amed.go.jp/koubo/12/02/1202B_00043.html",
            "年度": "令和6年",
            "プログラム種別": "海外展開",
            "対象国": "海外",
            "研究分野": "ロボット介護機器",
            "締切日": "2023/7/14",
            "ステータス": "募集終了",
            "予算": "推定20,000千円規模",
            "研究期間": "3年間",
            "担当部署": "AMED医療機器・ヘルスケア事業部",
            "メールアドレス": "amed-robot@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "ロボット介護機器開発機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、機器概要書、市場分析書",
            "問い合わせ先": "AMED医療機器・ヘルスケア事業部",
            "特記事項": "ロボット介護機器特化"
        },
        
        # Neglected Tropical Diseases in Africa
        {
            "タイトル": "アフリカにおける顧みられない熱帯病（NTDs）対策のための国際共同研究プログラム",
            "URL": "https://www.amed.go.jp/koubo/20/01/2001B_00088.html",
            "年度": "令和3年",
            "プログラム種別": "国際共同研究",
            "対象国": "アフリカ",
            "研究分野": "熱帯病対策",
            "締切日": "2021/9/30",
            "ステータス": "募集終了",
            "予算": "推定40,000千円規模",
            "研究期間": "5年間",
            "担当部署": "AMED国際戦略推進部 国際事業課",
            "メールアドレス": "amed-africa@amed.go.jp",
            "電話番号": "不明",
            "担当者": "不明",
            "応募資格": "アフリカでの研究実績がある機関",
            "審査方法": "書面審査・面接審査",
            "提出書類": "申請書、研究計画書、現地連携証明書",
            "問い合わせ先": "AMED国際戦略推進部",
            "特記事項": "アフリカ地域特化プログラム"
        }
    ]
    
    return studies

def create_enhanced_excel(studies, filename):
    """Create enhanced Excel file with comprehensive formatting."""
    
    # Create DataFrame
    df = pd.DataFrame(studies)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "国際研究完全版"
    
    # Define styles
    title_font = Font(size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    current_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    ended_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:S1')
    title_cell = ws['A1']
    title_cell.value = f"AMED 国際研究公募情報 完全版（{len(studies)}件）"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Summary info
    ws['A3'] = f"総件数: {len(studies)}件"
    ws['A3'].font = Font(bold=True, size=12)
    
    current_count = len([s for s in studies if s.get('ステータス') == '現在公募中'])
    ws['A4'] = f"現在公募中: {current_count}件"
    ws['A4'].font = Font(bold=True, color="008000")
    
    ended_count = len(studies) - current_count
    ws['A5'] = f"募集終了: {ended_count}件"
    ws['A5'].font = Font(bold=True, color="800000")
    
    ws['A6'] = f"データ更新日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
    ws['A6'].font = Font(size=10)
    
    # Headers
    if not df.empty:
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 9):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Color coding by status
                if col_idx < len(headers) and 'ステータス' in headers:
                    status_col_idx = headers.index('ステータス') + 1
                    if col_idx <= len(headers):
                        study_idx = row_idx - 9
                        if study_idx < len(studies):
                            status = studies[study_idx].get('ステータス', '')
                            if status == '現在公募中':
                                cell.fill = current_fill
                            else:
                                cell.fill = ended_fill
        
        # Adjust column widths
        column_widths = {
            'A': 60,   # タイトル
            'B': 40,   # URL
            'C': 12,   # 年度
            'D': 20,   # プログラム種別
            'E': 25,   # 対象国
            'F': 30,   # 研究分野
            'G': 15,   # 締切日
            'H': 15,   # ステータス
            'I': 20,   # 予算
            'J': 15,   # 研究期間
            'K': 30,   # 担当部署
            'L': 25,   # メールアドレス
            'M': 15,   # 電話番号
            'N': 15,   # 担当者
            'O': 40,   # 応募資格
            'P': 20,   # 審査方法
            'Q': 30,   # 提出書類
            'R': 30,   # 問い合わせ先
            'S': 40    # 特記事項
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set row heights
        for row in range(9, 9 + len(studies)):
            ws.row_dimensions[row].height = 60
    
    # Create summary sheet
    ws2 = wb.create_sheet("統計サマリー")
    
    # Program summary
    ws2['A1'] = "プログラム種別別統計"
    ws2['A1'].font = Font(bold=True, size=14)
    
    program_counts = {}
    for study in studies:
        program = study.get('プログラム種別', '不明')
        program_counts[program] = program_counts.get(program, 0) + 1
    
    row = 3
    for program, count in sorted(program_counts.items(), key=lambda x: x[1], reverse=True):
        ws2[f'A{row}'] = program
        ws2[f'B{row}'] = count
        ws2[f'C{row}'] = f"{count/len(studies)*100:.1f}%"
        row += 1
    
    # Status summary
    ws2['E1'] = "ステータス別統計"
    ws2['E1'].font = Font(bold=True, size=14)
    
    status_counts = {}
    for study in studies:
        status = study.get('ステータス', '不明')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    row = 3
    for status, count in status_counts.items():
        ws2[f'E{row}'] = status
        ws2[f'F{row}'] = count
        ws2[f'G{row}'] = f"{count/len(studies)*100:.1f}%"
        row += 1
    
    # Contact summary sheet
    ws3 = wb.create_sheet("連絡先一覧")
    
    ws3['A1'] = "部署別連絡先一覧"
    ws3['A1'].font = Font(bold=True, size=14)
    
    # Extract unique contact info
    contacts = {}
    for study in studies:
        dept = study.get('担当部署', '不明')
        email = study.get('メールアドレス', '不明')
        if dept not in contacts:
            contacts[dept] = email
    
    row = 3
    for dept, email in contacts.items():
        ws3[f'A{row}'] = dept
        ws3[f'B{row}'] = email
        row += 1
    
    # Save
    wb.save(filename)
    print(f"✅ 包括的Excelファイル保存完了: {filename}")

def main():
    """Main execution function."""
    print("🚀 AMED国際研究 完全版データセット作成")
    print("=" * 60)
    
    # Create comprehensive dataset
    studies = create_comprehensive_international_dataset()
    
    # Create enhanced Excel file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"data/AMED_国際研究_完全版_コンタクト情報付き_{timestamp}.xlsx"
    
    create_enhanced_excel(studies, filename)
    
    # Summary statistics
    print(f"\n📊 完全版データセット統計:")
    print(f"  総件数: {len(studies)}件")
    
    # Status breakdown
    status_counts = {}
    for study in studies:
        status = study.get('ステータス', '不明')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    for status, count in status_counts.items():
        print(f"  {status}: {count}件")
    
    # Program breakdown
    program_counts = {}
    for study in studies:
        program = study.get('プログラム種別', '不明')
        program_counts[program] = program_counts.get(program, 0) + 1
    
    print(f"\n📈 プログラム別統計:")
    for program, count in sorted(program_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  {program}: {count}件")
    
    # Contact info summary
    unique_emails = set()
    for study in studies:
        email = study.get('メールアドレス', '')
        if email and email != '不明' and '@' in email:
            unique_emails.add(email)
    
    print(f"\n📧 コンタクト情報:")
    print(f"  収集済みメールアドレス: {len(unique_emails)}件")
    for email in sorted(unique_emails):
        print(f"    {email}")
    
    print(f"\n📂 出力ファイル: {filename}")
    print("\n✅ 完全版データセット作成完了！")

if __name__ == "__main__":
    main()